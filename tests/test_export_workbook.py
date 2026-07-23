from io import BytesIO
import unittest
from unittest.mock import patch

from openpyxl import load_workbook

from app_core.export_service import build_workbook, workbook_bytes
from app_core.schemas import ExportRequest


def make_row(row_id, **overrides):
    row = {
        "id": row_id,
        "eval_mode": "full",
        "task_type": "T2I",
        "v_a": "A",
        "v_b": "B",
        "scene": "city",
        "filename": f"image-{row_id}.png",
        "overall": "tie",
        "aesthetic": "tie",
        "logic": "tie",
        "consistency": "tie",
        "fidelity": None,
        "worker": "alice",
        "timestamp": "2026-07-15T12:00:00+08:00",
        "duration_seconds": 5,
        "skipped": 0,
        "user_id": 1,
        "bad_case_tags_a": "[]",
        "bad_case_tags_b": "[]",
        "bad_case_categories_a": "[]",
        "bad_case_categories_b": "[]",
    }
    row.update(overrides)
    return row


def overall_metadata(sheet, label):
    for row in sheet.iter_rows(min_row=1, max_row=9):
        for index, cell in enumerate(row[:-1]):
            if cell.value == label:
                return row[index + 1]
    raise AssertionError(f"Overall 元数据缺少 {label}")


class ExportWorkbookTests(unittest.TestCase):
    def setUp(self):
        self.t2i_request = ExportRequest(
            task_type="T2I",
            v1="B",
            v2="A",
            scenes=["city", "zoo"],
            dimensions=["consistency", "aesthetic"],
            workers=["alice"],
            start_time="2026-07-15T00:00:00+08:00",
            end_time="2026-07-15T23:59:59+08:00",
            eval_modes=["full", "overall"],
            result_filter="a",
            bad_case_filter="all",
            include_images=False,
            include_bad_cases=True,
            include_duration=True,
        )
        self.t2i_rows = [
            make_row(1, scene="zoo", overall="tie", aesthetic="A", consistency="B", bad_case_tags_a='["模糊失焦"]'),
            make_row(2, eval_mode="overall", scene="city", overall="A", aesthetic=None, logic=None, consistency=None),
            make_row(3, scene="city", overall="B", aesthetic="B", consistency="A", worker="bob"),
            make_row(4, scene="city", overall="A", aesthetic="A", consistency="A", skipped=1),
        ]
        self.ti2i_request = ExportRequest(
            task_type="TI2I", v1="D", v2="E", dimensions=["fidelity"], include_bad_cases=False, include_duration=False
        )
        self.ti2i_rows = [
            make_row(1, task_type="TI2I", v_a="D", v_b="E", scene="portrait", fidelity="D")
        ]

    @patch("app_core.export_service.get_prompt_text", return_value="a red car")
    def test_t2i_workbook_has_overall_and_scene_ordered_sheets_with_metadata(self, _prompt):
        workbook = build_workbook(
            self.t2i_request, self.t2i_rows, generated_at="2026-07-15T10:00:00+08:00"
        )

        self.assertEqual(workbook.sheetnames, ["Overall", "city", "zoo"])
        overall = workbook["Overall"]
        self.assertEqual(overall["A1"].value, "评测结果导出")
        self.assertEqual(overall["B2"].value, "2026-07-15T10:00:00+08:00")
        self.assertEqual(overall["B3"].value, "T2I")
        self.assertEqual(overall["B4"].value, "A vs B")
        self.assertIn("city, zoo", [cell.value for row in overall.iter_rows(min_row=1, max_row=9) for cell in row])
        self.assertIn("一致性, 美学", [cell.value for row in overall.iter_rows(min_row=1, max_row=9) for cell in row])
        self.assertEqual(overall["A12"].value, "全部场景")

    @patch("app_core.export_service.get_prompt_text", return_value="a red car")
    def test_overall_and_detail_apply_result_filter_independently(self, _prompt):
        workbook = build_workbook(self.t2i_request, self.t2i_rows)
        overall = workbook["Overall"]
        detail = workbook["zoo"]

        self.assertEqual(overall["B12"].value, 1)
        self.assertEqual(detail.max_row, 3)
        headers = [cell.value for cell in detail[2]]
        self.assertEqual(detail.cell(3, headers.index("图片名") + 1).value, "image-1.png")
        self.assertEqual(detail.cell(3, headers.index("美学") + 1).value, "A")
        self.assertIsNone(detail.cell(3, headers.index("一致性") + 1).value)

    def test_overall_metadata_without_image_manifest_marks_images_unchecked_after_round_trip(self):
        overall = load_workbook(BytesIO(workbook_bytes(build_workbook(self.t2i_request, self.t2i_rows))))["Overall"]

        self.assertEqual(overall_metadata(overall, "最终评测记录数").value, 1)
        self.assertEqual(overall_metadata(overall, "最终评测记录数").data_type, "n")
        self.assertEqual(overall_metadata(overall, "缺失图片数量").value, "未检查")
        self.assertEqual(overall_metadata(overall, "缺失图片数量").data_type, "s")
        self.assertEqual(overall_metadata(overall, "图片状态").value, "未导出")
        self.assertEqual(overall["A11"].value, "场景")

    def test_overall_metadata_with_complete_manifest_marks_images_exported_after_round_trip(self):
        request = ExportRequest(task_type="TI2I", v1="D", v2="E", include_images=True)
        manifest = {
            ("portrait", "image-1.png"): {
                "D": {"status": "已导出"},
                "E": {"status": "已导出"},
                "ref": {"status": "已导出"},
            }
        }

        overall = load_workbook(
            BytesIO(workbook_bytes(build_workbook(request, self.ti2i_rows, image_manifest=manifest)))
        )["Overall"]

        self.assertEqual(overall_metadata(overall, "最终评测记录数").value, 1)
        self.assertEqual(overall_metadata(overall, "最终评测记录数").data_type, "n")
        self.assertEqual(overall_metadata(overall, "缺失图片数量").value, 0)
        self.assertEqual(overall_metadata(overall, "缺失图片数量").data_type, "n")
        self.assertEqual(overall_metadata(overall, "图片状态").value, "已导出")
        self.assertEqual(overall["A11"].value, "场景")

    def test_overall_metadata_counts_partial_missing_files_from_unique_manifest_after_round_trip(self):
        request = ExportRequest(task_type="TI2I", v1="D", v2="E", include_images=True)
        rows = self.ti2i_rows + [make_row(2, task_type="TI2I", v_a="D", v_b="E", scene="portrait", filename="image-1.png", fidelity="E", worker="bob")]
        manifest = {
            ("portrait", "image-1.png"): {
                "D": {"status": "文件不存在"},
                "E": {"status": "已导出"},
                "ref": {"status": "文件不存在"},
            }
        }

        overall = load_workbook(BytesIO(workbook_bytes(build_workbook(request, rows, image_manifest=manifest))))["Overall"]

        self.assertEqual(overall_metadata(overall, "最终评测记录数").value, 2)
        self.assertEqual(overall_metadata(overall, "最终评测记录数").data_type, "n")
        self.assertEqual(overall_metadata(overall, "缺失图片数量").value, 2)
        self.assertEqual(overall_metadata(overall, "缺失图片数量").data_type, "n")
        self.assertEqual(overall_metadata(overall, "图片状态").value, "部分缺失")
        self.assertEqual(overall["A11"].value, "场景")

    @patch("app_core.export_service.get_prompt_text", return_value="a red car")
    def test_scene_detail_contains_shared_result_group_and_required_fields(self, _prompt):
        sheet = build_workbook(self.t2i_request, self.t2i_rows)["zoo"]
        group_headers = [cell.value for cell in sheet[1]]
        headers = [cell.value for cell in sheet[2]]

        for header in ("任务类型", "模型 A", "模型 B", "场景", "图片名", "Prompt", "评测人", "评测模式", "评测时间（北京时间）"):
            self.assertIn(header, headers)
        self.assertEqual(group_headers.count("评测结果"), 1)
        self.assertIn("整体", headers)
        self.assertIn("美学", headers)
        self.assertIn("一致性", headers)
        self.assertIn("评测耗时（秒）", headers)
        self.assertIn("A 坏例标签", headers)
        self.assertEqual(sheet.cell(3, headers.index("Prompt") + 1).value, "a red car")
        self.assertEqual(sheet.cell(3, headers.index("A 图片路径") + 1).value, "")
        self.assertEqual(sheet.cell(3, headers.index("A 图片状态") + 1).value, "未导出")
        self.assertEqual(sheet.freeze_panes, "A3")
        self.assertEqual(sheet.auto_filter.ref, f"A2:{sheet.cell(2, sheet.max_column).coordinate[0:-1]}{sheet.max_row}")
        self.assertTrue(sheet.cell(3, headers.index("Prompt") + 1).alignment.wrap_text)
        result_range = next(
            cell_range
            for cell_range in sheet.merged_cells.ranges
            if sheet.cell(cell_range.min_row, cell_range.min_col).value == "评测结果"
        )
        self.assertEqual(result_range.max_col - result_range.min_col + 1, 3)

    @patch("app_core.export_service.get_prompt_text", return_value="portrait prompt")
    def test_ti2i_workbook_has_fidelity_reference_columns_and_round_trips(self, _prompt):
        rows = self.ti2i_rows + [
            make_row(2, task_type="TI2I", v_a="D", v_b="E", scene="portrait", fidelity="tie"),
            make_row(3, task_type="TI2I", v_a="D", v_b="E", scene="portrait", fidelity="E"),
        ]
        workbook = build_workbook(self.ti2i_request, rows)
        sheet = workbook["portrait"]
        headers = [cell.value for cell in sheet[2]]

        self.assertEqual([cell.value for cell in sheet[1]].count("评测结果"), 1)
        self.assertIn("整体", headers)
        self.assertIn("保真度", headers)
        self.assertIn("参考图路径", headers)
        self.assertIn("参考图状态", headers)
        self.assertNotIn("评测耗时（秒）", headers)
        self.assertNotIn("D 坏例标签", headers)
        self.assertEqual(sheet.cell(3, headers.index("参考图路径") + 1).value, "")
        self.assertEqual(sheet.cell(3, headers.index("参考图状态") + 1).value, "未导出")
        self.assertEqual(
            [sheet.cell(row, headers.index("保真度") + 1).value for row in range(3, 6)],
            ["D", "tie", "E"],
        )
        loaded = load_workbook(BytesIO(workbook_bytes(workbook)))
        self.assertEqual(loaded.sheetnames, ["Overall", "portrait"])

    @patch("app_core.export_service.get_prompt_text", return_value="portrait prompt")
    def test_overall_evaluated_row_is_written_to_existing_scene_detail_without_fabricating_dimensions(self, _prompt):
        request = ExportRequest(task_type="TI2I", v1="D", v2="E", dimensions=["fidelity"])
        rows = [make_row(
            1, task_type="TI2I", v_a="D", v_b="E", scene="portrait",
            eval_mode="overall", overall="D", fidelity=None,
        )]

        sheet = build_workbook(request, rows)["portrait"]
        headers = [cell.value for cell in sheet[2]]

        self.assertEqual(sheet.max_row, 3)
        self.assertEqual(sheet.cell(3, headers.index("图片名") + 1).value, "image-1.png")
        self.assertEqual(sheet.cell(3, headers.index("整体") + 1).value, "D")
        self.assertIsNone(sheet.cell(3, headers.index("保真度") + 1).value)

    @patch("app_core.export_service.get_prompt_text", return_value="portrait prompt")
    def test_overall_and_dimensions_use_single_raw_result_columns(self, _prompt):
        request = ExportRequest(
            task_type="TI2I", v1="D", v2="E", dimensions=["fidelity"],
            eval_modes=["full", "overall"],
        )
        rows = [
            make_row(1, task_type="TI2I", v_a="D", v_b="E", scene="portrait", eval_mode="overall", overall="D", fidelity=None),
            make_row(2, task_type="TI2I", v_a="D", v_b="E", scene="portrait", eval_mode="full", overall="tie", fidelity="E"),
        ]

        sheet = build_workbook(request, rows)["portrait"]
        groups = [cell.value for cell in sheet[1]]
        headers = [cell.value for cell in sheet[2]]

        self.assertEqual(groups.count("评测结果"), 1)
        self.assertEqual(headers.count("整体"), 1)
        self.assertEqual(headers.count("保真度"), 1)
        self.assertNotIn("D 胜", headers)
        self.assertNotIn("平局", headers)
        self.assertNotIn("E 胜", headers)
        self.assertEqual(
            [[sheet.cell(row, headers.index(name) + 1).value for name in ("整体", "保真度")] for row in (3, 4)],
            [["D", None], ["tie", "E"]],
        )

    @patch("app_core.export_service.get_prompt_text", return_value="portrait prompt")
    def test_full_only_export_omits_overall_result_column(self, _prompt):
        request = ExportRequest(
            task_type="TI2I", v1="D", v2="E", dimensions=["fidelity"], eval_modes=["full"],
        )
        rows = [make_row(1, task_type="TI2I", v_a="D", v_b="E", scene="portrait", overall="D", fidelity="tie")]

        sheet = build_workbook(request, rows)["portrait"]
        headers = [cell.value for cell in sheet[2]]

        self.assertNotIn("整体", headers)
        self.assertEqual(sheet.cell(3, headers.index("保真度") + 1).value, "tie")

    @patch("app_core.export_service.get_prompt_text", return_value="=SUM(1, 1)")
    def test_external_text_is_not_written_as_excel_formula(self, _prompt):
        request = ExportRequest(
            task_type="T2I",
            v1="=model-b",
            v2="+model-a",
            scenes=["-scene"],
            workers=["@worker"],
            dimensions=["aesthetic"],
        )
        rows = [
            make_row(
                1,
                v_a="+model-a",
                v_b="=model-b",
                overall="+model-a",
                aesthetic="=model-b",
                scene="-scene",
                filename="@image.png",
                worker="@worker",
            )
        ]

        loaded = load_workbook(BytesIO(workbook_bytes(build_workbook(request, rows))), data_only=False)

        for sheet in loaded.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    self.assertNotEqual(cell.data_type, "f", f"{sheet.title}!{cell.coordinate}")
        self.assertEqual(loaded["Overall"]["B4"].value, "'+model-a vs =model-b")
        self.assertEqual(loaded["Overall"]["B5"].value, "'-scene")
        detail = loaded["-scene"]
        headers = [cell.value for cell in detail[2]]
        self.assertEqual(detail.cell(3, headers.index("模型 A") + 1).value, "'+model-a")
        self.assertEqual(detail.cell(3, headers.index("模型 B") + 1).value, "'=model-b")
        self.assertEqual(detail.cell(3, headers.index("场景") + 1).value, "'-scene")
        self.assertEqual(detail.cell(3, headers.index("图片名") + 1).value, "'@image.png")
        self.assertEqual(detail.cell(3, headers.index("评测人") + 1).value, "'@worker")
        self.assertEqual(detail.cell(3, headers.index("Prompt") + 1).value, "'=SUM(1, 1)")

    @patch("app_core.export_service.get_prompt_text", return_value="prompt")
    def test_scene_sheet_titles_are_excel_safe_unique_and_limited(self, _prompt):
        request = ExportRequest(task_type="T2I", v1="A", v2="B", dimensions=["aesthetic"])
        rows = [
            make_row(1, scene="Overall", aesthetic="A"),
            make_row(2, scene="a:b", aesthetic="A"),
            make_row(3, scene="a/b", aesthetic="A"),
            make_row(4, scene="x" * 40, aesthetic="A"),
            make_row(5, scene="'quoted'", aesthetic="A"),
            make_row(6, scene="History", aesthetic="A"),
        ]

        workbook = build_workbook(request, rows)

        self.assertEqual(workbook.sheetnames[0], "Overall")
        self.assertEqual(len(workbook.sheetnames), 7)
        self.assertEqual(len({name.casefold() for name in workbook.sheetnames}), 7)
        self.assertTrue(all(len(name) <= 31 for name in workbook.sheetnames))
        self.assertTrue(all(not any(char in name for char in "[]:*?/\\") for name in workbook.sheetnames[1:]))
        self.assertTrue(all(not name.startswith("'") and not name.endswith("'") for name in workbook.sheetnames))
        self.assertNotIn("history", {name.casefold() for name in workbook.sheetnames})

    @patch("app_core.export_service.get_prompt_text", return_value="cached prompt")
    def test_build_workbook_caches_prompt_by_task_scene_and_filename(self, prompt_lookup):
        request = ExportRequest(task_type="T2I", v1="A", v2="B", dimensions=["aesthetic", "logic"])
        rows = [
            make_row(1, filename="shared.png", aesthetic="A"),
            make_row(2, filename="shared.png", aesthetic="B", worker="bob"),
        ]

        build_workbook(request, rows)

        prompt_lookup.assert_called_once_with("T2I", "city", "shared.png")

    def test_overall_statistics_cover_empty_winners_ties_and_single_side_bad_cases(self):
        request = ExportRequest(task_type="T2I", v1="A", v2="B")
        cases = [
            ("empty", [], [0, 0, 0, 0, 0, 0, 0, "-", "-", 0, 0, 0, 0]),
            ("a_wins", [make_row(1, overall="A")], [1, 1, 1, 0, 0, 0, 0, "∞", 0, 0, 0, 0, 0]),
            ("b_wins", [make_row(1, overall="B")], [1, 0, 0, 0, 0, 1, 1, 0, "∞", 0, 0, 0, 0]),
            (
                "ties",
                [make_row(1, overall="tie"), make_row(2, overall="tie")],
                [2, 0, 0, 2, 1, 0, 0, 1, 1, 0, 0, 0, 0],
            ),
            (
                "a_bad_only",
                [make_row(1, overall="tie", bad_case_tags_a='["模糊失焦"]')],
                [1, 0, 0, 1, 1, 0, 0, 1, 1, 1, 1, 0, 0],
            ),
        ]

        for name, rows, expected in cases:
            with self.subTest(name=name):
                workbook = build_workbook(request, rows)
                overall = workbook["Overall"]
                values = [overall.cell(12, column).value for column in range(2, 15)]
                self.assertEqual(values, expected)
                for column in (4, 6, 8, 12, 14):
                    self.assertEqual(overall.cell(12, column).number_format, "0.0%")

                loaded = load_workbook(BytesIO(workbook_bytes(workbook)), data_only=False)["Overall"]
                for column in (2, 3, 4, 5, 6, 7, 8, 11, 12, 13, 14):
                    self.assertEqual(loaded.cell(12, column).data_type, "n")
                for column in (9, 10):
                    expected_type = "s" if isinstance(overall.cell(12, column).value, str) else "n"
                    self.assertEqual(loaded.cell(12, column).data_type, expected_type)


if __name__ == "__main__":
    unittest.main()
