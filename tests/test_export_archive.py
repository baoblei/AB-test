import tempfile
import unittest
import zipfile
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app_core.errors import AppError, ValidationError
from app_core.schemas import ExportRequest
from app_core.storage import get_ref_image_path, get_result_image_path


def make_row(row_id=1, **overrides):
    row = {
        "id": row_id, "eval_mode": "full", "task_type": "TI2I", "v_a": "D", "v_b": "E",
        "scene": "portrait", "filename": "scene1.jpg", "overall": "D", "aesthetic": "D",
        "logic": "D", "consistency": "D", "fidelity": "D", "worker": "alice",
        "timestamp": "2026-07-15T12:00:00+08:00", "duration_seconds": 3, "skipped": 0,
        "user_id": 1, "bad_case_tags_a": "[]", "bad_case_tags_b": "[]",
        "bad_case_categories_a": "[]", "bad_case_categories_b": "[]",
    }
    row.update(overrides)
    return row


class ExportArchiveTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        root = Path(self.temp_dir.name)
        self.results = root / "results"
        self.refs = root / "refs"
        for model in ("D", "E"):
            directory = self.results / model / "portrait"
            directory.mkdir(parents=True)
            (directory / "scene1.jpg").write_bytes(model.encode())
        (self.refs / "portrait").mkdir(parents=True)
        (self.refs / "portrait" / "scene1.jpg").write_bytes(b"ref")
        self.request = ExportRequest(task_type="TI2I", v1="D", v2="E", dimensions=["fidelity"], include_images=True)

    def tearDown(self):
        self.temp_dir.cleanup()

    def result_resolver(self, task_type, model, scene, filename):
        path = self.results / model / scene / filename
        return str(path) if path.is_file() else None

    def ref_resolver(self, task_type, scene, filename):
        path = self.refs / scene / filename
        return str(path) if path.is_file() else None

    def test_storage_image_paths_reject_unsafe_components(self):
        for value in ("", ".", "..", "a/b", "a\\b"):
            with self.subTest(value=value):
                with self.assertRaises(AppError):
                    get_result_image_path("T2I", value, "scene", "image.png")
                with self.assertRaises(AppError):
                    get_ref_image_path("TI2I", "scene", value)

    def test_result_image_path_rejects_symlink_components_and_marks_manifest_missing(self):
        from app_core.export_service import build_image_manifest

        request = ExportRequest(task_type="T2I", v1="D", v2="E", include_images=True)
        row = make_row(task_type="T2I")
        for component in ("root", "version", "scene", "filename"):
            with self.subTest(component=component):
                base = Path(self.temp_dir.name) / f"result-symlink-{component}"
                base.mkdir()
                root = base / "root"
                if component == "root":
                    target = base / "outside-root"
                    (target / "D" / "portrait").mkdir(parents=True)
                    (target / "D" / "portrait" / "scene1.jpg").write_bytes(b"outside")
                    root.symlink_to(target, target_is_directory=True)
                elif component == "version":
                    root.mkdir()
                    target = base / "outside-version"
                    (target / "portrait").mkdir(parents=True)
                    (target / "portrait" / "scene1.jpg").write_bytes(b"outside")
                    (root / "D").symlink_to(target, target_is_directory=True)
                elif component == "scene":
                    (root / "D").mkdir(parents=True)
                    target = base / "outside-scene"
                    target.mkdir()
                    (target / "scene1.jpg").write_bytes(b"outside")
                    (root / "D" / "portrait").symlink_to(target, target_is_directory=True)
                else:
                    (root / "D" / "portrait").mkdir(parents=True)
                    target = base / "outside-file.jpg"
                    target.write_bytes(b"outside")
                    (root / "D" / "portrait" / "scene1.jpg").symlink_to(target)

                with patch("app_core.storage.get_result_roots", return_value=[str(root)]):
                    self.assertIsNone(get_result_image_path("T2I", "D", "portrait", "scene1.jpg"))
                    manifest = build_image_manifest(
                        request,
                        [row],
                        result_path_resolver=get_result_image_path,
                    )
                self.assertEqual(manifest[("portrait", "scene1.jpg")]["D"]["status"], "文件不存在")
                self.assertIsNone(manifest[("portrait", "scene1.jpg")]["D"]["source_path"])

        root = Path(self.temp_dir.name) / "result-resolution-error"
        (root / "D" / "portrait").mkdir(parents=True)
        (root / "D" / "portrait" / "scene1.jpg").write_bytes(b"inside")
        outside = Path(self.temp_dir.name) / "resolved-outside.jpg"
        outside.write_bytes(b"outside")
        resolved_root = root.resolve()
        resolved_outside = outside.resolve()
        with patch("app_core.storage.get_result_roots", return_value=[str(root)]):
            with patch.object(Path, "resolve", side_effect=[resolved_root, resolved_outside]):
                self.assertIsNone(get_result_image_path("T2I", "D", "portrait", "scene1.jpg"))
        with patch("app_core.storage.get_result_roots", return_value=[str(root)]):
            with patch.object(Path, "resolve", side_effect=OSError("resolution failed")):
                self.assertIsNone(get_result_image_path("T2I", "D", "portrait", "scene1.jpg"))

    def test_ref_image_path_rejects_symlink_components_and_marks_manifest_missing(self):
        from app_core.export_service import build_image_manifest

        for component in ("root", "scene", "filename"):
            with self.subTest(component=component):
                base = Path(self.temp_dir.name) / f"ref-symlink-{component}"
                base.mkdir()
                root = base / "root"
                if component == "root":
                    target = base / "outside-root"
                    (target / "portrait").mkdir(parents=True)
                    (target / "portrait" / "scene1.jpg").write_bytes(b"outside")
                    root.symlink_to(target, target_is_directory=True)
                elif component == "scene":
                    root.mkdir()
                    target = base / "outside-scene"
                    target.mkdir()
                    (target / "scene1.jpg").write_bytes(b"outside")
                    (root / "portrait").symlink_to(target, target_is_directory=True)
                else:
                    (root / "portrait").mkdir(parents=True)
                    target = base / "outside-file.jpg"
                    target.write_bytes(b"outside")
                    (root / "portrait" / "scene1.jpg").symlink_to(target)

                with patch("app_core.storage.get_ref_root", return_value=str(root)):
                    with patch("app_core.storage.REF_IMAGE_DIR", str(base / "missing-legacy")):
                        self.assertIsNone(get_ref_image_path("TI2I", "portrait", "scene1.jpg"))
                        manifest = build_image_manifest(
                            self.request,
                            [make_row()],
                            result_path_resolver=lambda *_args: None,
                            ref_path_resolver=get_ref_image_path,
                        )
                self.assertEqual(manifest[("portrait", "scene1.jpg")]["ref"]["status"], "文件不存在")
                self.assertIsNone(manifest[("portrait", "scene1.jpg")]["ref"]["source_path"])

        root = Path(self.temp_dir.name) / "ref-resolution-error"
        (root / "portrait").mkdir(parents=True)
        (root / "portrait" / "scene1.jpg").write_bytes(b"inside")
        with patch("app_core.storage.get_ref_root", return_value=str(root)):
            with patch("app_core.storage.REF_IMAGE_DIR", str(Path(self.temp_dir.name) / "missing-legacy")):
                with patch.object(Path, "resolve", side_effect=OSError("resolution failed")):
                    self.assertIsNone(get_ref_image_path("TI2I", "portrait", "scene1.jpg"))

    def test_archive_uses_scene_model_ref_layout_and_deduplicates_rows(self):
        from app_core.export_service import build_archive

        archive_path = Path(self.temp_dir.name) / "archive.zip"
        artifact = build_archive(
            self.request,
            workbook_data=b"xlsx",
            selected_rows=[make_row(), make_row(2, worker="bob")],
            result_path_resolver=self.result_resolver,
            ref_path_resolver=self.ref_resolver,
            archive_path=str(archive_path),
        )
        with zipfile.ZipFile(artifact, "r") as archive:
            self.assertEqual(
                sorted(archive.namelist()),
                ["images/portrait/D/scene1.jpg", "images/portrait/E/scene1.jpg", "images/portrait/ref/scene1.jpg", "评测结果.xlsx"],
            )

    def test_result_image_path_and_archive_fall_back_to_legacy_file_after_configured_scene(self):
        from app_core.export_service import build_archive, build_image_manifest

        configured_root = Path(self.temp_dir.name) / "configured"
        legacy_root = Path(self.temp_dir.name) / "legacy"
        for root in (configured_root, legacy_root):
            for model in ("D", "E"):
                (root / model / "portrait").mkdir(parents=True)
        for model in ("D", "E"):
            (legacy_root / model / "portrait" / "scene1.jpg").write_bytes(model.encode())
        request = ExportRequest(task_type="T2I", v1="D", v2="E", include_images=True)
        rows = [make_row(task_type="T2I")]

        with patch("app_core.storage.get_result_roots", return_value=[str(configured_root), str(legacy_root)]):
            self.assertEqual(
                get_result_image_path("T2I", "D", "portrait", "scene1.jpg"),
                str(legacy_root / "D" / "portrait" / "scene1.jpg"),
            )
            manifest = build_image_manifest(request, rows, result_path_resolver=get_result_image_path)

        archive_path = Path(self.temp_dir.name) / "legacy-fallback.zip"
        build_archive(request, b"xlsx", rows, archive_path=str(archive_path), image_manifest=manifest)
        with zipfile.ZipFile(archive_path, "r") as archive:
            self.assertEqual(
                sorted(archive.namelist()),
                ["images/portrait/D/scene1.jpg", "images/portrait/E/scene1.jpg", "评测结果.xlsx"],
            )

    def test_missing_image_is_not_written_and_workbook_marks_manifest_status(self):
        from app_core.export_service import build_image_manifest, build_workbook, workbook_bytes

        manifest = build_image_manifest(self.request, [make_row()], lambda *args: None, lambda *args: None)
        self.assertEqual(manifest[("portrait", "scene1.jpg")]["D"]["status"], "文件不存在")
        workbook = build_workbook(self.request, [make_row()], image_manifest=manifest)
        sheet = load_workbook(BytesIO(workbook_bytes(workbook)))["portrait"]
        headers = [cell.value for cell in sheet[2]]
        self.assertEqual(sheet.cell(3, headers.index("D 图片路径") + 1).value, "images/portrait/D/scene1.jpg")
        self.assertEqual(sheet.cell(3, headers.index("D 图片状态") + 1).value, "文件不存在")

    def test_create_artifact_rejects_empty_overall_and_cleans_failed_temp_directory(self):
        from app_core import export_service

        with patch.object(export_service, "fetch_base_rows", return_value=[]):
            with self.assertRaisesRegex(AppError, "没有符合条件的评测记录"):
                export_service.create_export_artifact(self.request)

        temp_root = Path(self.temp_dir.name) / "export-root"
        temp_root.mkdir()
        with patch.object(export_service, "fetch_base_rows", return_value=[make_row()]):
            with patch.object(export_service.tempfile, "mkdtemp", return_value=str(temp_root / "artifact")):
                with patch.object(export_service, "build_workbook", side_effect=OSError("write failed")):
                    with self.assertRaises(OSError):
                        export_service.create_export_artifact(self.request)
        self.assertFalse((temp_root / "artifact").exists())

    def test_workbook_artifact_handles_invalid_timestamps_according_to_time_bounds(self):
        from app_core import export_service

        rows = [
            make_row(1, filename="canonical.jpg", timestamp="2026-07-15T12:00:00+08:00"),
            make_row(2, filename="null.jpg", timestamp=None),
            make_row(3, filename="legacy.jpg", timestamp="2026-07-15 11:00:00"),
        ]
        cases = [
            (ExportRequest(task_type="TI2I", v1="D", v2="E", dimensions=["fidelity"]), 3),
            (
                ExportRequest(
                    task_type="TI2I",
                    v1="D",
                    v2="E",
                    dimensions=["fidelity"],
                    start_time="2026-07-15T00:00:00+08:00",
                ),
                1,
            ),
        ]

        for request, expected_count in cases:
            with self.subTest(start_time=request.start_time):
                with patch.object(export_service, "fetch_base_rows", return_value=rows):
                    artifact = export_service.create_export_artifact(request)
                try:
                    workbook = load_workbook(artifact.path)
                    self.assertEqual(workbook["Overall"]["B9"].value, expected_count)
                    self.assertEqual(workbook["portrait"].max_row - 2, expected_count)
                    workbook.close()
                finally:
                    Path(artifact.path).unlink(missing_ok=True)
                    Path(artifact.cleanup_dir).rmdir()

    def test_ti2i_image_export_rejects_ref_models_before_preview_or_artifact_creation(self):
        from app_core import export_service

        temp_root = Path(self.temp_dir.name) / "ref-collision"
        temp_root.mkdir()
        for model in ("ref", "REF"):
            with self.subTest(model=model):
                request = ExportRequest(task_type="TI2I", v1=model, v2="Z", include_images=True)
                with self.assertRaisesRegex(ValidationError, "TI2I 导出图片时模型名称 ref 与参考图目录冲突"):
                    export_service.validate_export_request(request)
                with patch.object(export_service, "fetch_base_rows") as fetch_rows:
                    with self.assertRaisesRegex(ValidationError, "TI2I 导出图片时模型名称 ref 与参考图目录冲突"):
                        export_service.preview_export(request)
                fetch_rows.assert_not_called()
                with patch.object(export_service.tempfile, "mkdtemp") as make_temp_dir:
                    with self.assertRaisesRegex(ValidationError, "TI2I 导出图片时模型名称 ref 与参考图目录冲突"):
                        export_service.create_export_artifact(request)
                make_temp_dir.assert_not_called()
        self.assertEqual(list(temp_root.iterdir()), [])

    def test_direct_manifest_rejects_ti2i_ref_model_without_image_export_flag(self):
        from app_core.export_service import build_image_manifest

        for model in ("ref", "REF"):
            with self.subTest(model=model):
                request = ExportRequest(task_type="TI2I", v1=model, v2="Z", include_images=False)
                with self.assertRaisesRegex(
                    ValidationError, "TI2I 导出图片时模型名称 ref 与参考图目录冲突"
                ) as context:
                    build_image_manifest(request, [make_row(v_a=model, v_b="Z")])
                self.assertEqual(context.exception.status_code, 422)

    def test_direct_archive_rejects_ti2i_ref_model_before_tempfile_with_prebuilt_manifest(self):
        from app_core import export_service

        manifest = {("portrait", "scene1.jpg"): {}}
        for model in ("ref", "REF"):
            with self.subTest(model=model):
                request = ExportRequest(task_type="TI2I", v1=model, v2="Z", include_images=False)
                with patch.object(export_service.tempfile, "mkstemp") as make_temp_file:
                    with patch.object(export_service.zipfile, "ZipFile") as create_archive:
                        with self.assertRaisesRegex(
                            ValidationError, "TI2I 导出图片时模型名称 ref 与参考图目录冲突"
                        ) as context:
                            export_service.build_archive(
                                request,
                                workbook_data=b"xlsx",
                                selected_rows=[make_row(v_a=model, v_b="Z")],
                                image_manifest=manifest,
                            )
                self.assertEqual(context.exception.status_code, 422)
                make_temp_file.assert_not_called()
                create_archive.assert_not_called()

    def test_ti2i_ref_collision_uses_trimmed_canonical_model_names(self):
        from app_core import export_service

        for model in (" ref", "ref ", " REF "):
            with self.subTest(model=model):
                request = ExportRequest(task_type="TI2I", v1=model, v2="Z", include_images=True)
                with self.assertRaisesRegex(ValidationError, "TI2I 导出图片时模型名称 ref 与参考图目录冲突"):
                    export_service.validate_export_request(request)

        with self.assertRaisesRegex(ValidationError, "模型必须不同"):
            export_service.validate_export_request(
                ExportRequest(task_type="T2I", v1="model", v2=" model ", include_images=True)
            )

    def test_archive_does_not_follow_source_replaced_after_manifest_creation(self):
        from app_core.export_service import build_archive, build_image_manifest

        request = ExportRequest(task_type="T2I", v1="D", v2="E", include_images=True)
        source = self.results / "D" / "portrait" / "scene1.jpg"
        outside = Path(self.temp_dir.name) / "outside.txt"
        outside.write_bytes(b"outside-secret")
        manifest = build_image_manifest(
            request,
            [make_row(task_type="T2I")],
            result_path_resolver=self.result_resolver,
        )
        source.unlink()
        source.symlink_to(outside)

        archive_path = Path(self.temp_dir.name) / "replaced-source.zip"
        build_archive(
            request,
            b"xlsx",
            [make_row(task_type="T2I")],
            archive_path=str(archive_path),
            image_manifest=manifest,
        )

        with zipfile.ZipFile(archive_path) as archive:
            self.assertNotIn("images/portrait/D/scene1.jpg", archive.namelist())
            self.assertNotIn(b"outside-secret", [archive.read(name) for name in archive.namelist()])

    def test_ti2i_ref_model_allows_pure_xlsx_export(self):
        from app_core import export_service

        request = ExportRequest(task_type="TI2I", v1="ref", v2="Z", include_images=False)
        artifact_dir = Path(self.temp_dir.name) / "xlsx-artifact"

        def make_temp_dir(**_kwargs):
            artifact_dir.mkdir()
            return str(artifact_dir)

        with patch.object(export_service, "fetch_base_rows", return_value=[make_row(v_a="Z", v_b="ref")]):
            with patch.object(export_service.tempfile, "mkdtemp", side_effect=make_temp_dir):
                artifact = export_service.create_export_artifact(request)

        self.assertEqual(artifact.media_type, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertTrue(Path(artifact.path).is_file())
        Path(artifact.path).unlink()
        artifact_dir.rmdir()

    def test_t2i_ref_model_allows_image_manifest(self):
        from app_core.export_service import build_image_manifest

        request = ExportRequest(task_type="T2I", v1="ref", v2="Z", include_images=True)
        manifest = build_image_manifest(request, [make_row(task_type="T2I", v_a="Z", v_b="ref")], lambda *args: None)

        self.assertIn("ref", manifest[("portrait", "scene1.jpg")])


class ExportApiTests(unittest.TestCase):
    def test_new_export_routes_preserve_legacy_get_and_return_file_response(self):
        import main
        from app_core.export_service import ExportArtifact

        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "评测结果.xlsx"
            path.write_bytes(b"xlsx")
            artifact = ExportArtifact(str(path), "评测结果.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", temp_dir)
            with patch.object(main, "export_results", return_value={"legacy": True}):
                with patch.object(main, "get_export_options", return_value={"total": 1}) as options:
                    with patch.object(main, "preview_export", return_value={"overall": 1}) as preview:
                        with patch.object(main, "create_export_artifact", return_value=artifact) as create:
                            client = TestClient(main.app)
                            self.assertEqual(client.get("/api/export", params={"format": "json"}).status_code, 200)
                            self.assertEqual(
                                client.get(
                                    "/api/export_options", params={"task_type": "T2I", "v1": "A", "v2": "B"}
                                ).status_code,
                                401,
                            )
                            self.assertEqual(
                                client.post(
                                    "/api/export/preview", json={"task_type": "T2I", "v1": "A", "v2": "B"}
                                ).status_code,
                                401,
                            )
                            self.assertEqual(
                                client.post(
                                    "/api/export", json={"task_type": "T2I", "v1": "A", "v2": "B"}
                                ).status_code,
                                401,
                            )

                            main.app.dependency_overrides[main.require_login] = lambda: {
                                "id": 1,
                                "username": "reviewer",
                            }
                            try:
                                self.assertEqual(
                                    client.get(
                                        "/api/export_options",
                                        params={"task_type": "T2I", "v1": "A", "v2": "B"},
                                    ).json(),
                                    {"total": 1},
                                )
                                self.assertEqual(
                                    client.post(
                                        "/api/export/preview",
                                        json={"task_type": "T2I", "v1": "A", "v2": "B"},
                                    ).json(),
                                    {"overall": 1},
                                )
                                response = client.post(
                                    "/api/export", json={"task_type": "T2I", "v1": "A", "v2": "B"}
                                )
                            finally:
                                main.app.dependency_overrides.pop(main.require_login, None)
            options.assert_called_once_with("T2I", "A", "B")
            preview.assert_called_once()
            create.assert_called_once()
            self.assertEqual(response.headers["content-type"], artifact.media_type)
            self.assertIn("filename*=utf-8''", response.headers["content-disposition"])
            self.assertFalse(path.exists())

    def test_export_request_validation_returns_422_without_changing_generic_app_errors(self):
        import main

        client = TestClient(main.app)
        main.app.dependency_overrides[main.require_login] = lambda: {"id": 1, "username": "reviewer"}
        try:
            semantic_response = client.post(
                "/api/export/preview", json={"task_type": "invalid", "v1": "A", "v2": "B"}
            )
            pydantic_response = client.post(
                "/api/export/preview", json={"task_type": "T2I", "v1": "A"}
            )
            with patch.object(main, "preview_export", side_effect=AppError("普通业务错误")):
                generic_response = client.post(
                    "/api/export/preview", json={"task_type": "T2I", "v1": "A", "v2": "B"}
                )
        finally:
            main.app.dependency_overrides.pop(main.require_login, None)

        self.assertEqual(semantic_response.status_code, 422)
        self.assertEqual(pydantic_response.status_code, 422)
        self.assertEqual(generic_response.status_code, 400)


if __name__ == "__main__":
    unittest.main()
