import os
import re
import shutil
import tempfile
import zipfile
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .bad_cases import safe_load_json_list
from .config import (
    DIM_LABELS,
    TASK_CONFIGS,
    dim_payload,
    get_task_config,
    is_video_task,
    normalize_task_type,
)
from .database import connect
from .errors import AppError, ValidationError
from .schemas import ExportRequest
from .storage import (
    copy_regular_file_without_symlinks,
    get_regular_file_identity,
    get_prompt_text,
    get_ref_image_path,
    get_result_image_path,
    validate_storage_component,
)
from .thumbnail_service import get_image_thumbnail
from .time_utils import is_canonical_beijing_iso, now_beijing_iso


VALID_RESULT_FILTERS = {"all", "a", "tie_bad", "tie_good", "b"}
VALID_BAD_CASE_FILTERS = {"all", "with", "without"}
VALID_EVAL_MODES = {"full", "overall", "selected"}

INVALID_SHEET_TITLE_CHARS = re.compile(r"[\[\]:*?/\\]")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAPPED_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
EXCEL_FORMULA_PREFIXES = ("=", "+", "-", "@")
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
ZIP_MEDIA_TYPE = "application/zip"
TI2I_REF_MODEL_COLLISION_ERROR = "TI2I 导出图片时模型名称 ref 与参考图目录冲突"


@dataclass
class ExportArtifact:
    path: str
    filename: str
    media_type: str
    cleanup_dir: str


def canonical_models(request: ExportRequest) -> tuple[str, str]:
    return canonical_model_pair(request.v1, request.v2)


def canonical_model_pair(v1: str, v2: str) -> tuple[str, str]:
    normalized_v1 = v1.strip() if isinstance(v1, str) else ""
    normalized_v2 = v2.strip() if isinstance(v2, str) else ""
    if not normalized_v1 or not normalized_v2:
        raise ValidationError("模型名称不能为空")
    if normalized_v1 == normalized_v2:
        raise ValidationError("模型必须不同")
    return tuple(sorted((normalized_v1, normalized_v2)))


def _validate_ti2i_ref_model_collision(task_type: str, v_a: str, v_b: str) -> None:
    config = get_task_config(task_type)
    if not config["upload_has_ref"] or not any(
        model.casefold() == "ref" for model in (v_a, v_b)
    ):
        return
    if task_type == "TI2I":
        raise ValidationError(TI2I_REF_MODEL_COLLISION_ERROR)
    raise ValidationError(f"{task_type} 导出媒体时模型名称 ref 与参考图目录冲突")


def validate_export_request(request: ExportRequest) -> tuple[str, str, str]:
    task_type = normalize_task_type(request.task_type)
    if task_type not in TASK_CONFIGS:
        raise ValidationError("无效任务类型")

    v_a, v_b = canonical_models(request)
    if request.include_images:
        _validate_ti2i_ref_model_collision(task_type, v_a, v_b)
    config = get_task_config(task_type)
    video_task = config["media_type"] == "video"
    valid_dimensions = set(
        config["dashboard_dims"] if video_task else config["eval_dims"]
    )
    if any(dimension not in valid_dimensions for dimension in request.dimensions):
        raise ValidationError("无效导出维度")
    if request.result_filter not in VALID_RESULT_FILTERS:
        raise ValidationError("无效结果筛选")
    if request.bad_case_filter not in VALID_BAD_CASE_FILTERS:
        raise ValidationError("无效坏例筛选")
    if any(mode not in VALID_EVAL_MODES for mode in request.eval_modes):
        raise ValidationError("无效评测模式")
    if video_task and request.eval_modes != ["selected"]:
        raise ValidationError("视频导出仅支持自定义维度评测模式")
    if not video_task and "selected" in request.eval_modes:
        raise ValidationError("图片导出不支持自定义维度评测模式")
    if request.start_time and not is_canonical_beijing_iso(request.start_time):
        raise ValidationError("导出时间必须为北京时间 ISO 格式")
    if request.end_time and not is_canonical_beijing_iso(request.end_time):
        raise ValidationError("导出时间必须为北京时间 ISO 格式")
    if request.start_time and request.end_time and request.start_time > request.end_time:
        raise ValidationError("开始时间不能晚于结束时间")
    return task_type, v_a, v_b


def expected_result(request: ExportRequest, v_a: Optional[str] = None, v_b: Optional[str] = None) -> Optional[str]:
    if v_a is None or v_b is None:
        v_a, v_b = canonical_models(request)
    return {
        "all": None,
        "a": v_a,
        "tie_bad": "tie_bad",
        "tie_good": "tie_good",
        "b": v_b,
    }[request.result_filter]


def row_has_bad_case(row) -> bool:
    return bool(safe_load_json_list(row["bad_case_tags_a"]) or safe_load_json_list(row["bad_case_tags_b"]))


def _is_canonical_export_timestamp(value) -> bool:
    return isinstance(value, str) and is_canonical_beijing_iso(value)


def filter_rows(rows: Iterable, request: ExportRequest, dimension: str) -> list:
    task_type, v_a, v_b = validate_export_request(request)
    configured_dimensions = get_task_config(task_type)["dashboard_dims"]
    if dimension not in configured_dimensions:
        raise ValidationError("无效导出维度")

    wanted_result = expected_result(request, v_a, v_b)
    result = []
    for row in rows:
        if (
            row["task_type"] != task_type
            or row["v_a"] != v_a
            or row["v_b"] != v_b
            or row["skipped"] != 0
        ):
            continue
        mode = row["eval_mode"] or "full"
        if request.scenes and row["scene"] not in request.scenes:
            continue
        if request.workers and row["worker"] not in request.workers:
            continue
        if request.eval_modes and mode not in request.eval_modes:
            continue
        if request.start_time or request.end_time:
            timestamp = row["timestamp"]
            if not _is_canonical_export_timestamp(timestamp):
                continue
            if request.start_time and timestamp < request.start_time:
                continue
            if request.end_time and timestamp > request.end_time:
                continue
        has_bad_case = row_has_bad_case(row)
        if request.bad_case_filter == "with" and not has_bad_case:
            continue
        if request.bad_case_filter == "without" and has_bad_case:
            continue
        if mode == "selected":
            if row[dimension] is None:
                continue
        elif dimension != "overall" and (mode != "full" or row[dimension] is None):
            continue
        if wanted_result and row[dimension] != wanted_result:
            continue
        result.append(row)
    return result


def fetch_base_rows(task_type: str, v_a: str, v_b: str) -> list:
    conn = connect(row_factory=True)
    try:
        rows = conn.execute(
            """
            SELECT * FROM results_log
            WHERE task_type=? AND v_a=? AND v_b=? AND skipped=0
            """,
            (task_type, v_a, v_b),
        ).fetchall()
    finally:
        conn.close()
    return rows


def get_export_options(task_type: str, v1: str, v2: str) -> dict:
    normalized_task_type = normalize_task_type(task_type)
    if normalized_task_type not in TASK_CONFIGS:
        raise ValidationError("无效任务类型")
    v_a, v_b = canonical_model_pair(v1, v2)
    rows = fetch_base_rows(normalized_task_type, v_a, v_b)
    timestamps = [row["timestamp"] for row in rows if _is_canonical_export_timestamp(row["timestamp"])]
    config = TASK_CONFIGS[normalized_task_type]
    return {
        "task_type": normalized_task_type,
        "v_a": v_a,
        "v_b": v_b,
        "scenes": sorted({row["scene"] for row in rows}),
        "workers": sorted({row["worker"] for row in rows}),
        "dimensions": dim_payload(
            config["dashboard_dims"]
            if config["media_type"] == "video"
            else config["eval_dims"]
        ),
        "media_type": config["media_type"],
        "min_time": min(timestamps) if timestamps else None,
        "max_time": max(timestamps) if timestamps else None,
        "total": len(rows),
    }


def preview_export(request: ExportRequest) -> dict:
    task_type, v_a, v_b = validate_export_request(request)
    rows = fetch_base_rows(task_type, v_a, v_b)
    overall_rows = (
        filter_rows(rows, request, "overall")
        if not is_video_task(task_type) or "overall" in request.dimensions
        else []
    )
    dimension_rows = {
        dimension: filter_rows(rows, request, dimension)
        for dimension in request.dimensions
        if dimension != "overall"
    }
    selected_rows = overall_rows + [row for rows_for_dimension in dimension_rows.values() for row in rows_for_dimension]
    unique_images = {(row["scene"], row["filename"]) for row in selected_rows}
    return {
        "overall": len(overall_rows),
        "dimensions": {dimension: len(items) for dimension, items in dimension_rows.items()},
        "unique_images": len(unique_images),
    }


def suppression_ratio(numerator: int, denominator: int):
    if denominator == 0:
        return "∞" if numerator else "-"
    return round(numerator / denominator, 2)


def summarize_overall(rows, v_a: str, v_b: str) -> dict:
    total = len(rows)
    a_wins = sum(row["overall"] == v_a for row in rows)
    tie_bad = sum(row["overall"] == "tie_bad" for row in rows)
    tie_good = sum(row["overall"] == "tie_good" for row in rows)
    ties = tie_bad + tie_good
    b_wins = sum(row["overall"] == v_b for row in rows)
    bad_a = sum(bool(safe_load_json_list(row["bad_case_tags_a"])) for row in rows)
    bad_b = sum(bool(safe_load_json_list(row["bad_case_tags_b"])) for row in rows)
    return {
        "total": total,
        "a_wins": a_wins,
        "tie_bad": tie_bad,
        "tie_good": tie_good,
        "ties": ties,
        "b_wins": b_wins,
        "a_rate": a_wins / total if total else 0,
        "tie_bad_rate": tie_bad / total if total else 0,
        "tie_good_rate": tie_good / total if total else 0,
        "b_rate": b_wins / total if total else 0,
        "a_suppression": suppression_ratio(a_wins + ties, b_wins + ties),
        "b_suppression": suppression_ratio(b_wins + ties, a_wins + ties),
        "bad_a": bad_a,
        "bad_b": bad_b,
        "bad_a_rate": bad_a / total if total else 0,
        "bad_b_rate": bad_b / total if total else 0,
    }


def excel_safe_text(value):
    if isinstance(value, str) and value.startswith(EXCEL_FORMULA_PREFIXES):
        return f"'{value}"
    return value


def _selected_values(values: list[str], labels: Optional[dict] = None) -> str:
    if not values:
        return "全部"
    return ", ".join(labels.get(value, value) if labels else value for value in values)


def _flag_value(value: bool) -> str:
    return "是" if value else "否"


def _image_manifest_summary(image_manifest: Optional[dict]) -> tuple[str, object]:
    if image_manifest is None:
        return "未导出", "未检查"
    missing_count = sum(
        image["status"] == "文件不存在"
        for images in image_manifest.values()
        for image in images.values()
    )
    return ("部分缺失" if missing_count else "已导出"), missing_count


def _style_header(sheet, row: int, columns: int) -> None:
    for column in range(1, columns + 1):
        cell = sheet.cell(row, column)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _fit_columns(sheet, wrapped_headers: Optional[set[str]] = None, header_row: int = 1) -> None:
    wrapped_headers = {excel_safe_text(header) for header in wrapped_headers or set()}
    for column in range(1, sheet.max_column + 1):
        header = sheet.cell(header_row, column).value
        max_length = max(len(str(sheet.cell(row, column).value or "")) for row in range(1, sheet.max_row + 1))
        width = min(max(max_length + 2, 10), 60)
        sheet.column_dimensions[get_column_letter(column)].width = width
        if header in wrapped_headers:
            for row in range(header_row + 1, sheet.max_row + 1):
                sheet.cell(row, column).alignment = WRAPPED_ALIGNMENT


def _write_overall_metadata(
    sheet,
    request: ExportRequest,
    task_type: str,
    v_a: str,
    v_b: str,
    generated_at: str,
    final_record_count: int,
    image_manifest: Optional[dict],
) -> None:
    sheet["A1"] = "评测结果导出"
    sheet["A1"].font = Font(bold=True, size=14)
    image_status, missing_image_count = _image_manifest_summary(image_manifest)
    video_task = is_video_task(task_type)
    media_label = "视频" if video_task else "图片"
    metadata = [
        ("生成时间", generated_at),
        ("任务类型", task_type),
        ("模型对", f"{v_a} vs {v_b}"),
        ("场景", _selected_values(request.scenes), "维度", _selected_values(request.dimensions, DIM_LABELS), "评测人", _selected_values(request.workers)),
        ("开始时间", request.start_time or "不限", "结束时间", request.end_time or "不限", "评测模式", _selected_values(request.eval_modes)),
        ("结果筛选", request.result_filter, "坏例筛选", request.bad_case_filter, f"导出{media_label}", _flag_value(request.include_images)),
        ("包含坏例", _flag_value(request.include_bad_cases), "包含耗时", _flag_value(request.include_duration), f"{media_label}状态", image_status),
        ("最终评测记录数", final_record_count, f"缺失{media_label}数量", missing_image_count),
    ]
    for row_number, values in enumerate(metadata, start=2):
        for column, value in enumerate(values, start=1):
            sheet.cell(row_number, column, excel_safe_text(value))
        for column in range(1, len(values) + 1, 2):
            sheet.cell(row_number, column).font = Font(bold=True)


def _summary_values(scene: str, rows: list, v_a: str, v_b: str) -> list:
    stats = summarize_overall(rows, v_a, v_b)
    return [
        scene,
        stats["total"],
        stats["a_wins"],
        stats["a_rate"],
        stats["tie_bad"],
        stats["tie_bad_rate"],
        stats["tie_good"],
        stats["tie_good_rate"],
        stats["b_wins"],
        stats["b_rate"],
        stats["a_suppression"],
        stats["b_suppression"],
        stats["bad_a"],
        stats["bad_a_rate"],
        stats["bad_b"],
        stats["bad_b_rate"],
    ]


def _write_overall_sheet(
    sheet,
    request: ExportRequest,
    rows: list,
    task_type: str,
    v_a: str,
    v_b: str,
    generated_at: str,
    image_manifest: Optional[dict],
) -> None:
    _write_overall_metadata(
        sheet,
        request,
        task_type,
        v_a,
        v_b,
        generated_at,
        len(rows),
        image_manifest,
    )
    headers = [
        "场景", "总数", f"{v_a} 胜数", f"{v_a} 胜率", "一样差数", "一样差率", "一样好数", "一样好率",
        f"{v_b} 胜数", f"{v_b} 胜率", f"{v_a} 抑制比", f"{v_b} 抑制比", f"{v_a} 坏例数", f"{v_a} 坏例率", f"{v_b} 坏例数", f"{v_b} 坏例率",
    ]
    for column, header in enumerate(headers, start=1):
        sheet.cell(11, column, excel_safe_text(header))
    for column, value in enumerate(_summary_values("全部场景", rows, v_a, v_b), start=1):
        sheet.cell(12, column, excel_safe_text(value) if column == 1 else value)
    for row_number, scene in enumerate(sorted({row["scene"] for row in rows}), start=13):
        scene_rows = [row for row in rows if row["scene"] == scene]
        for column, value in enumerate(_summary_values(scene, scene_rows, v_a, v_b), start=1):
            sheet.cell(row_number, column, excel_safe_text(value) if column == 1 else value)
    _style_header(sheet, 11, len(headers))
    for row in range(12, sheet.max_row + 1):
        for column in (4, 6, 8, 10, 14, 16):
            sheet.cell(row, column).number_format = "0.0%"
    sheet.auto_filter.ref = f"A11:{get_column_letter(len(headers))}{sheet.max_row}"
    sheet.freeze_panes = "A12"
    for column in range(1, sheet.max_column + 1):
        max_length = max(len(str(sheet.cell(row, column).value or "")) for row in range(1, sheet.max_row + 1))
        sheet.column_dimensions[get_column_letter(column)].width = min(max(max_length + 2, 10), 30)


def _selected_result_dimensions(
    dimensions: list[str], request: ExportRequest, task_type: str
) -> list[str]:
    if is_video_task(task_type):
        requested = set(dimensions)
        return [
            dimension
            for dimension in get_task_config(task_type)["dashboard_dims"]
            if dimension in requested
        ]
    result = ["overall"] if "overall" in request.eval_modes else []
    result.extend(dimensions)
    return result


def _scene_detail_groups(
    dimensions: list[str], request: ExportRequest, task_type: str, v_a: str, v_b: str
) -> list[tuple[str, list[str]]]:
    video_task = is_video_task(task_type)
    filename_label = "文件名" if video_task else "图片名"
    sample_headers = [
        "任务类型", "模型 A", "模型 B", "场景", filename_label, "Prompt", "评测人", "评测模式", "评测时间（北京时间）",
    ]
    if request.include_duration:
        sample_headers.append("评测耗时（秒）")
    groups = [("样本信息", sample_headers)]
    result_dimensions = _selected_result_dimensions(dimensions, request, task_type)
    if result_dimensions:
        groups.append(("评测结果", [DIM_LABELS[dimension] for dimension in result_dimensions]))
    if video_task:
        image_headers = [
            f"{v_a} 视频路径", f"{v_a} 视频状态", f"{v_a} 首帧路径",
            f"{v_b} 视频路径", f"{v_b} 视频状态", f"{v_b} 首帧路径",
        ]
        group_label = "视频信息"
    else:
        image_headers = [f"{v_a} 图片路径", f"{v_a} 图片状态", f"{v_b} 图片路径", f"{v_b} 图片状态"]
        group_label = "图片信息"
    if get_task_config(task_type)["upload_has_ref"]:
        image_headers.extend(["参考图路径", "参考图状态"])
    groups.append((group_label, image_headers))
    if request.include_bad_cases:
        groups.append(
            ("坏例信息", [f"{v_a} 坏例标签", f"{v_a} 坏例类别", f"{v_b} 坏例标签", f"{v_b} 坏例类别"])
        )
    return groups


def _scene_detail_values(
    row,
    dimensions: list[str],
    matching_row_ids: dict[str, set[int]],
    request: ExportRequest,
    task_type: str,
    v_a: str,
    v_b: str,
    prompt_cache: dict[tuple[str, str, str], str],
    image_manifest: Optional[dict],
) -> list:
    prompt_key = (task_type, row["scene"], row["filename"])
    if prompt_key not in prompt_cache:
        prompt_cache[prompt_key] = get_prompt_text(*prompt_key)
    image_info = (image_manifest or {}).get((row["scene"], row["filename"]), {})
    a_image = image_info.get(v_a, {})
    b_image = image_info.get(v_b, {})
    row_id = row["id"]
    values = [
        task_type, v_a, v_b, row["scene"], row["filename"], prompt_cache[prompt_key],
        row["worker"], row["eval_mode"] or "full", row["timestamp"],
    ]
    if request.include_duration:
        values.append(row["duration_seconds"])
    result_labels = {"tie_bad": "一样差", "tie_good": "一样好"}
    for dimension in _selected_result_dimensions(dimensions, request, task_type):
        result = row[dimension] if row_id in matching_row_ids[dimension] else None
        values.append(result_labels.get(result, result))
    if is_video_task(task_type):
        values.extend(
            [
                a_image.get("path", ""), a_image.get("status", "未导出"), a_image.get("poster_path", ""),
                b_image.get("path", ""), b_image.get("status", "未导出"), b_image.get("poster_path", ""),
            ]
        )
    else:
        values.extend(
            [
                a_image.get("path", ""), a_image.get("status", "未导出"),
                b_image.get("path", ""), b_image.get("status", "未导出"),
            ]
        )
    if get_task_config(task_type)["upload_has_ref"]:
        ref_image = image_info.get("ref", {})
        values.extend([ref_image.get("path", ""), ref_image.get("status", "未导出")])
    if request.include_bad_cases:
        values.extend(
            [
                ", ".join(safe_load_json_list(row["bad_case_tags_a"])),
                ", ".join(safe_load_json_list(row["bad_case_categories_a"])),
                ", ".join(safe_load_json_list(row["bad_case_tags_b"])),
                ", ".join(safe_load_json_list(row["bad_case_categories_b"])),
            ]
        )
    return values


def _write_scene_detail_sheet(
    sheet,
    dimensions: list[str],
    matching_row_ids: dict[str, set[int]],
    request: ExportRequest,
    rows: list,
    task_type: str,
    v_a: str,
    v_b: str,
    prompt_cache: dict[tuple[str, str, str], str],
    image_manifest: Optional[dict],
) -> None:
    groups = _scene_detail_groups(dimensions, request, task_type, v_a, v_b)
    group_row = []
    header_row = []
    column = 1
    merged_ranges = []
    for group_name, headers in groups:
        group_row.extend([excel_safe_text(group_name)] + [None] * (len(headers) - 1))
        header_row.extend(excel_safe_text(header) for header in headers)
        if len(headers) > 1:
            merged_ranges.append((column, column + len(headers) - 1))
        column += len(headers)
    sheet.append(group_row)
    sheet.append(header_row)
    for start_column, end_column in merged_ranges:
        sheet.merge_cells(start_row=1, start_column=start_column, end_row=1, end_column=end_column)
    for row in rows:
        values = _scene_detail_values(
            row, dimensions, matching_row_ids, request, task_type, v_a, v_b, prompt_cache, image_manifest
        )
        sheet.append([excel_safe_text(value) for value in values])
    _style_header(sheet, 1, len(header_row))
    _style_header(sheet, 2, len(header_row))
    sheet.auto_filter.ref = f"A2:{get_column_letter(len(header_row))}{max(sheet.max_row, 2)}"
    sheet.freeze_panes = "A3"
    _fit_columns(
        sheet,
        {"Prompt", f"{v_a} 坏例标签", f"{v_a} 坏例类别", f"{v_b} 坏例标签", f"{v_b} 坏例类别"},
        header_row=2,
    )


def _scene_sheet_title(scene: str, existing_titles: set[str]) -> str:
    cleaned = INVALID_SHEET_TITLE_CHARS.sub("_", str(scene))
    cleaned = "".join(character if ord(character) >= 32 else "_" for character in cleaned)
    cleaned = cleaned.strip().strip("'").strip() or "场景"
    cleaned = cleaned[:31]
    candidate = cleaned
    suffix_number = 2
    while candidate.casefold() in existing_titles:
        suffix = f" ({suffix_number})"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        suffix_number += 1
    existing_titles.add(candidate.casefold())
    return candidate


def build_workbook(
    request: ExportRequest,
    rows: Iterable,
    generated_at: Optional[str] = None,
    image_manifest: Optional[dict] = None,
) -> Workbook:
    task_type, v_a, v_b = validate_export_request(request)
    base_rows = list(rows)
    workbook = Workbook()
    overall = workbook.active
    overall.title = "Overall"
    overall_rows = (
        filter_rows(base_rows, request, "overall")
        if not is_video_task(task_type) or "overall" in request.dimensions
        else []
    )
    _write_overall_sheet(
        overall,
        request,
        overall_rows,
        task_type,
        v_a,
        v_b,
        generated_at or now_beijing_iso(),
        image_manifest,
    )

    requested_dimensions = set(request.dimensions)
    configured_dimensions = (
        get_task_config(task_type)["dashboard_dims"]
        if is_video_task(task_type)
        else get_task_config(task_type)["eval_dims"]
    )
    dimensions = [
        dimension
        for dimension in configured_dimensions
        if dimension in requested_dimensions
    ]
    dimension_rows = {
        dimension: filter_rows(base_rows, request, dimension)
        for dimension in dimensions
        if dimension != "overall"
    }
    matching_row_ids = {"overall": {row["id"] for row in overall_rows}}
    matching_row_ids.update({
        dimension: {row["id"] for row in rows_for_dimension}
        for dimension, rows_for_dimension in dimension_rows.items()
    })
    detail_row_ids = set().union(*matching_row_ids.values()) if matching_row_ids else set()
    detail_scenes = {row["scene"] for row in overall_rows}
    detail_scenes.update(row["scene"] for rows_for_dimension in dimension_rows.values() for row in rows_for_dimension)
    prompt_cache = {}
    existing_titles = {"overall", "history"}
    for scene in sorted(detail_scenes):
        sheet = workbook.create_sheet(_scene_sheet_title(scene, existing_titles))
        scene_rows = [row for row in base_rows if row["id"] in detail_row_ids and row["scene"] == scene]
        _write_scene_detail_sheet(
            sheet,
            dimensions,
            matching_row_ids,
            request,
            scene_rows,
            task_type,
            v_a,
            v_b,
            prompt_cache,
            image_manifest,
        )
    return workbook


def workbook_bytes(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _archive_path(
    task_type: str,
    scene: str,
    model: str,
    filename: str,
    source_path: Optional[str] = None,
) -> str:
    if not is_video_task(task_type):
        return "/".join(("images", scene, model, filename))
    source_extension = Path(source_path or filename).suffix.lower()
    media_filename = f"{Path(filename).stem}{source_extension}"
    return "/".join(("videos", scene, model, media_filename))


def _poster_archive_path(scene: str, model: str, filename: str) -> str:
    return "/".join(("posters", scene, model, f"{Path(filename).stem}.webp"))


def _default_poster_path_resolver(
    task_type: str, model: str, scene: str, filename: str
) -> str:
    return get_image_thumbnail(
        "result", task_type, scene, filename, model=model
    )


def build_image_manifest(
    request: ExportRequest,
    selected_rows: Iterable,
    result_path_resolver=get_result_image_path,
    ref_path_resolver=get_ref_image_path,
    poster_path_resolver=_default_poster_path_resolver,
) -> dict:
    task_type, v_a, v_b = validate_export_request(request)
    _validate_ti2i_ref_model_collision(task_type, v_a, v_b)
    config = get_task_config(task_type)
    video_task = config["media_type"] == "video"
    manifest = {}
    for row in selected_rows:
        scene = validate_storage_component(row["scene"], "场景")
        filename = validate_storage_component(row["filename"], "图片名")
        key = (scene, filename)
        if key in manifest:
            continue
        entry = {}
        for model in (v_a, v_b):
            model = validate_storage_component(model, "模型")
            source_path = result_path_resolver(task_type, model, scene, filename)
            source_identity = get_regular_file_identity(source_path) if source_path else None
            entry[model] = {
                "path": _archive_path(task_type, scene, model, filename, source_path),
                "status": "已导出" if source_identity else "文件不存在",
                "source_path": source_path if source_identity else None,
                "_source_identity": source_identity,
            }
            if video_task:
                poster_source_path = None
                if source_identity:
                    try:
                        poster_source_path = poster_path_resolver(
                            task_type, model, scene, filename
                        )
                    except (AppError, OSError):
                        poster_source_path = None
                poster_identity = (
                    get_regular_file_identity(poster_source_path)
                    if poster_source_path
                    else None
                )
                entry[model].update(
                    {
                        "poster_path": _poster_archive_path(scene, model, filename),
                        "poster_status": "已导出" if poster_identity else "文件不存在",
                        "poster_source_path": poster_source_path if poster_identity else None,
                        "_poster_source_identity": poster_identity,
                    }
                )
        if config["upload_has_ref"]:
            source_path = ref_path_resolver(task_type, scene, filename)
            source_identity = get_regular_file_identity(source_path) if source_path else None
            reference_filename = filename
            if video_task and source_path:
                reference_filename = f"{Path(filename).stem}{Path(source_path).suffix.lower()}"
            entry["ref"] = {
                "path": "/".join(("images", scene, "ref", reference_filename)),
                "status": "已导出" if source_identity else "文件不存在",
                "source_path": source_path if source_identity else None,
                "_source_identity": source_identity,
            }
        manifest[key] = entry
    return manifest


def snapshot_image_manifest(image_manifest: dict, snapshot_dir: str) -> dict:
    os.makedirs(snapshot_dir, exist_ok=True)
    snapshot = {}
    image_index = 0
    for key, images in image_manifest.items():
        snapshot_images = {}
        for model, image in images.items():
            snapshot_image = dict(image)
            source_path = image.get("source_path")
            if source_path:
                destination = os.path.join(snapshot_dir, f"{image_index}.bin")
                image_index += 1
                if copy_regular_file_without_symlinks(
                    source_path,
                    destination,
                    image.get("_source_identity"),
                ):
                    snapshot_image["source_path"] = destination
                    snapshot_image["status"] = "已导出"
                    snapshot_image["_source_identity"] = get_regular_file_identity(destination)
                else:
                    snapshot_image["source_path"] = None
                    snapshot_image["status"] = "文件不存在"
                    snapshot_image["_source_identity"] = None
            poster_source_path = image.get("poster_source_path")
            if poster_source_path:
                destination = os.path.join(snapshot_dir, f"{image_index}.bin")
                image_index += 1
                if copy_regular_file_without_symlinks(
                    poster_source_path,
                    destination,
                    image.get("_poster_source_identity"),
                ):
                    snapshot_image["poster_source_path"] = destination
                    snapshot_image["poster_status"] = "已导出"
                    snapshot_image["_poster_source_identity"] = get_regular_file_identity(destination)
                else:
                    snapshot_image["poster_source_path"] = None
                    snapshot_image["poster_status"] = "文件不存在"
                    snapshot_image["_poster_source_identity"] = None
            snapshot_images[model] = snapshot_image
        snapshot[key] = snapshot_images
    return snapshot


def build_archive(
    request: ExportRequest,
    workbook_data: bytes,
    selected_rows: Iterable,
    result_path_resolver=get_result_image_path,
    ref_path_resolver=get_ref_image_path,
    archive_path: Optional[str] = None,
    image_manifest: Optional[dict] = None,
) -> str:
    task_type, v_a, v_b = validate_export_request(request)
    _validate_ti2i_ref_model_collision(task_type, v_a, v_b)
    manifest = image_manifest or build_image_manifest(request, selected_rows, result_path_resolver, ref_path_resolver)
    with tempfile.TemporaryDirectory(prefix="ab-test-export-snapshot-") as snapshot_dir:
        stable_manifest = snapshot_image_manifest(manifest, snapshot_dir)
        if archive_path is None:
            descriptor, archive_path = tempfile.mkstemp(prefix="ab-test-export-", suffix=".zip")
            os.close(descriptor)
        with zipfile.ZipFile(archive_path, "w", zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("评测结果.xlsx", workbook_data)
            for entry in stable_manifest.values():
                for model in (v_a, v_b):
                    image = entry[model]
                    if image["source_path"]:
                        archive.write(image["source_path"], image["path"])
                    if image.get("poster_source_path"):
                        archive.write(image["poster_source_path"], image["poster_path"])
                if get_task_config(task_type)["upload_has_ref"] and entry["ref"]["source_path"]:
                    archive.write(entry["ref"]["source_path"], entry["ref"]["path"])
    return archive_path


def create_export_artifact(request: ExportRequest) -> ExportArtifact:
    task_type, v_a, v_b = validate_export_request(request)
    rows = fetch_base_rows(task_type, v_a, v_b)
    overall_rows = (
        filter_rows(rows, request, "overall")
        if not is_video_task(task_type) or "overall" in request.dimensions
        else []
    )
    dimension_rows = [
        filter_rows(rows, request, dimension)
        for dimension in request.dimensions
        if dimension != "overall"
    ]
    selected_rows = overall_rows + [row for items in dimension_rows for row in items]
    if not selected_rows:
        raise AppError("当前筛选条件下没有符合条件的评测记录，无法生成导出文件")
    cleanup_dir = tempfile.mkdtemp(prefix="ab-test-export-")
    try:
        manifest = build_image_manifest(request, selected_rows) if request.include_images else None
        if manifest is not None:
            manifest = snapshot_image_manifest(manifest, os.path.join(cleanup_dir, "image-snapshots"))
        workbook_data = workbook_bytes(build_workbook(request, rows, image_manifest=manifest))
        generated_at = now_beijing_iso().replace(":", "-").replace("+", "_")
        task_label = validate_storage_component(task_type, "任务类型")
        a_label = validate_storage_component(v_a, "模型")
        b_label = validate_storage_component(v_b, "模型")
        stem = f"评测导出_{task_label}_{a_label}_vs_{b_label}_{generated_at}"
        if not request.include_images:
            path = os.path.join(cleanup_dir, f"{stem}.xlsx")
            with open(path, "wb") as output:
                output.write(workbook_data)
            return ExportArtifact(path, os.path.basename(path), XLSX_MEDIA_TYPE, cleanup_dir)
        path = os.path.join(cleanup_dir, f"{stem}.zip")
        build_archive(request, workbook_data, selected_rows, archive_path=path, image_manifest=manifest)
        return ExportArtifact(path, os.path.basename(path), ZIP_MEDIA_TYPE, cleanup_dir)
    except Exception:
        shutil.rmtree(cleanup_dir, ignore_errors=True)
        raise
